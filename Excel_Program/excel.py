import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# load work book
wb = xl.load_workbook('Excel_Program/transactions.xlsx')

# load one of the sheet
sheet = wb['Sheet1']

# refer in rows and column
# cell = sheet['a1'] :- instead of writing this we can write :-

# cell = sheet.cell(1,1)
# print(sheet.max_rows) to get no of rows

for row in range(2,sheet.max_row + 1):
    cell = sheet.cell(row,3)
    correct_price = cell.value * 0.9
    correct_price_cell = sheet.cell(row,4)
    correct_price_cell.value = correct_price

# to create a chat
val = Reference(
          sheet,
          min_row = 2,
          max_row = sheet.max_row,
          min_col = 4,
          max_col = 4)

chart = BarChart()
chart.add_data(val)
sheet.add_chart(chart,'e2')

# save the workbook in seprate file
wb.save('Excel_Program/corrected_book.xlsx')